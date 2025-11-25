"""
Servicio para generación de reportes PDF y Excel
"""
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from app.models.student import Student
from app.models.professor import Professor
from app.services.statistics_service import StatisticsService
from app import db

class ReportService:
    
    @staticmethod
    def generate_students_pdf():
        """Genera PDF con todos los estudiantes y sus estadísticas"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        title = Paragraph("REPORTE DE ESTUDIANTES", title_style)
        subtitle = Paragraph(
            f"Valle Grande de Cañete - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            styles['Normal']
        )
        
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Obtener estudiantes
        students = db.session.query(Student).filter(Student.status == 'A').all()
        
        # Crear tabla
        data = [['Código', 'Nombre', 'Carrera', 'Asistencia', 'Estado']]
        
        for student in students:
            stats = StatisticsService.calculate_student_attendance(student.id)
            attendance_pct = f"{stats['attendance_percentage']}%"
            risk = stats['risk_level']
            
            risk_text = {
                'NORMAL': 'Normal',
                'ATENCION': 'Atención',
                'EN_RIESGO': 'En Riesgo',
                'CRITICO': 'CRÍTICO'
            }.get(risk, 'Normal')
            
            data.append([
                student.student_code,
                student.full_name,
                student.career,
                attendance_pct,
                risk_text
            ])
        
        table = Table(data, colWidths=[1.2*inch, 2.5*inch, 2*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_student_pdf(student_id):
        """Genera PDF individual de un estudiante"""
        student = db.session.query(Student).get(student_id)
        if not student:
            return None
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph(f"REPORTE DE ESTUDIANTE", styles['Heading1'])
        subtitle = Paragraph(
            f"{student.full_name} - {student.student_code}",
            styles['Heading2']
        )
        
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Información personal
        info_data = [
            ['Campo', 'Valor'],
            ['Código', student.student_code],
            ['Nombre Completo', student.full_name],
            ['Email', student.email],
            ['Carrera', student.career],
            ['Semestre', student.semester],
            ['DNI', student.dni or 'N/A'],
            ['Teléfono', student.phone or 'N/A'],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas de asistencia
        stats = StatisticsService.calculate_student_attendance(student_id)
        
        stats_title = Paragraph("ESTADÍSTICAS DE ASISTENCIA", styles['Heading2'])
        elements.append(stats_title)
        elements.append(Spacer(1, 0.2*inch))
        
        stats_data = [
            ['Métrica', 'Valor'],
            ['Total de Clases', str(stats['total_classes'])],
            ['Presentes', str(stats['present'])],
            ['Ausentes', str(stats['absent'])],
            ['Justificadas', str(stats['justified'])],
            ['Tardanzas', str(stats['late'])],
            ['% Asistencia', f"{stats['attendance_percentage']}%"],
            ['% Inasistencias', f"{stats['absence_percentage']}%"],
            ['Nivel de Riesgo', stats['risk_level']],
        ]
        
        stats_table = Table(stats_data, colWidths=[2*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(stats_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_students_excel():
        """Genera Excel con todos los estudiantes"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estudiantes"
        
        # Estilos
        header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['Código', 'Nombre Completo', 'Email', 'Carrera', 'Semestre', 
                   'Total Clases', 'Presentes', 'Ausentes', '% Asistencia', 'Estado']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        students = db.session.query(Student).filter(Student.status == 'A').all()
        
        for row, student in enumerate(students, 2):
            stats = StatisticsService.calculate_student_attendance(student.id)
            
            ws.cell(row=row, column=1, value=student.student_code).border = border
            ws.cell(row=row, column=2, value=student.full_name).border = border
            ws.cell(row=row, column=3, value=student.email).border = border
            ws.cell(row=row, column=4, value=student.career).border = border
            ws.cell(row=row, column=5, value=student.semester).border = border
            ws.cell(row=row, column=6, value=stats['total_classes']).border = border
            ws.cell(row=row, column=7, value=stats['present']).border = border
            ws.cell(row=row, column=8, value=stats['absent']).border = border
            ws.cell(row=row, column=9, value=f"{stats['attendance_percentage']}%").border = border
            ws.cell(row=row, column=10, value=stats['risk_level']).border = border
            
            # Color según riesgo
            risk_colors = {
                'CRITICO': 'ffcccc',
                'EN_RIESGO': 'ffe6cc',
                'ATENCION': 'ffffcc',
                'NORMAL': 'ccffcc'
            }
            risk_color = risk_colors.get(stats['risk_level'], 'ffffff')
            
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color=risk_color,
                    end_color=risk_color,
                    fill_type="solid"
                )
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def generate_professors_pdf():
        """Genera PDF con todos los profesores"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title = Paragraph("REPORTE DE PROFESORES", styles['Heading1'])
        subtitle = Paragraph(
            f"Valle Grande de Cañete - {datetime.now().strftime('%d/%m/%Y')}",
            styles['Normal']
        )
        
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))
        
        # Obtener profesores
        professors = db.session.query(Professor).filter(Professor.status == 'A').all()
        
        # Crear tabla
        data = [['Código', 'Nombre', 'Especialización', 'Departamento', 'Rol']]
        
        for prof in professors:
            data.append([
                prof.professor_code,
                prof.full_name,
                prof.specialization or 'N/A',
                prof.department or 'N/A',
                prof.role
            ])
        
        table = Table(data, colWidths=[1.2*inch, 2.5*inch, 2*inch, 1.5*inch, 1.3*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        return buffer
